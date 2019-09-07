
# coding = utf-8

import re
from pymongo import MongoClient
import csv
import os
import time
import datetime
from openpyxl import load_workbook
from setting import oriDataBusiness, oriDataCourt, oriDataOperation, oriDataPath, rootPath
import pandas as pd
from 新冒烟指数_广告 import GetAdData

if not os.path.exists(oriDataBusiness):
    os.makedirs(oriDataBusiness)
if not os.path.exists(oriDataCourt):
    os.makedirs(oriDataCourt)
if not os.path.exists(oriDataOperation):
    os.makedirs(oriDataOperation)

client = MongoClient('192.168.1.110', 17017)
db = client['download']


class GetAllBusinessCSV:

    def __init__(self, file):

        self.companys = file

    # 基本信息
    def getJibenxinxi(self):

        col = db['corp_lkl_faceinfo']

        f_jibenxinxi = open(os.path.join(oriDataBusiness, '基本信息.csv'), 'w', encoding='utf-8', newline='')
        headers_jibenxinxi = ['企业名称', '法定代表人', '统一社会信用代码', '工商注册号', '注册资金（万元）',
                              '开业时间', '经营期限自', '企业类型', '经营状态', '登记机关', '注册地址',
                              '经营范围', '实收资本', '简称', '核准时间', '省份简称', '邮箱', '行业', '是否是小微企业（0不是 1是）',
                              '法人id', '组织机构代码', '电话号码', '人员规模', '纳税人识别号', '类型', '网址', '注销原因',
                              '币种', '曾用名', '吊销日期', '吊销原因', '经营期限至', '主管机构名称', '注销日期']
        writer_jibenxinxi = csv.DictWriter(f_jibenxinxi, headers_jibenxinxi)
        writer_jibenxinxi.writeheader()
        start = time.time()
        data_jibenxinxi = col.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$sort': {'storage_time': -1}},
            {'$group': {'_id': "$entName",
                        '企业名称': {'$first': '$entName'},
                        '法定代表人': {'$first': '$frName'},
                        '统一社会信用代码': {'$first': '$authorityCode'},
                        '工商注册号': {'$first': '$regNo'},
                        '注册资金（万元）': {'$first': '$regCap'},
                        '开业时间': {'$first': '$esDate'},
                        '经营期限自': {'$first': '$opFrom'},
                        '企业类型': {'$first': '$entType'},
                        '经营状态': {'$first': '$entStatus'},
                        '登记机关': {'$first': '$regOrg'},
                        '注册地址': {'$first': '$dom'},
                        '经营范围': {'$first': '$opScope'},
                        '实收资本': {'$first': '$recCap'},
                        '简称': {'$first': '$alias'},
                        '核准时间': {'$first': '$approvedTime'},
                        '省份简称': {'$first': '$base'},
                        '邮箱': {'$first': '$email'},
                        '行业': {'$first': '$industry'},
                        '是否是小微企业（0不是 1是）': {'$first': '$isMicroEnt'},
                        '法人id': {'$first': '$legalPersonId'},
                        '组织机构代码': {'$first': '$orgNo'},
                        '电话号码': {'$first': '$phone'},
                        '人员规模': {'$first': '$staffNum'},
                        '纳税人识别号': {'$first': '$taxNumber'},
                        '类型': {'$first': '$type'},
                        '网址': {'$first': '$websiteList'},
                        '注销原因': {'$first': '$cancelReason'},
                        '币种': {'$first': '$regCapCur'},
                        '曾用名': {'$first': '$historyNames'},
                        '吊销日期': {'$first': '$revDate'},
                        '吊销原因': {'$first': '$revokeReason'},
                        '经营期限至': {'$first': '$opTo'},
                        '主管机构名称': {'$first': '$authorityName'},
                        '注销日期': {'$first': '$canDate'},
                        }
             },
            {'$project': {
                '_id': 0,
            }}
        ])

        count = 0
        for data in data_jibenxinxi:
            # 此处转换时间戳  # aa = int(str(tt)[:10])
            # bb = tt/1000  该方法弃用，位数不一定
            # data['storage_time'] = datetime.datetime.fromtimestamp(data['storage_time']/1000).strftime("%Y-%m-%d %H:%M:%S")
            try:
                # if data['开业时间']:
                #     data['开业时间'] = datetime.datetime.fromtimestamp(int(str(data['开业时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                # if data['经营期限自']:
                #     data['经营期限自'] = datetime.datetime.fromtimestamp(int(str(data['经营期限自'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                # if data['经营期限至']:
                #     data['经营期限至'] = datetime.datetime.fromtimestamp(int(str(data['经营期限至'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                writer_jibenxinxi.writerow(data)
                count += 1
                # print(data)
            except Exception as e:
                print(e)
                print(data)
        print(count)
        print('基本信息 Over!' + str(time.time()-start))
        f_jibenxinxi.close()

    # 对外投资
    def getDuiwaitouzi(self):

        db = client['download']['corp_lkl_entinv']

        f_duiwaitouzi = open(os.path.join(oriDataBusiness, '对外投资.csv'), 'w', encoding='utf-8', newline='')
        headers_duiwaitouzi = ['企业名称', '投资企业名称', '公司类型', '经营范围', '投资占比', '企业状态', '法人',
                               '投资金额', '行业', '注册资金', '省份简称', '统一社会信用代码', '简称', '被投资企业开业日期']
        writer_duiwaitouzi = csv.DictWriter(f_duiwaitouzi, headers_duiwaitouzi)
        writer_duiwaitouzi.writeheader()

        start = time.time()
        IDs_duiwaitouzi = db.aggregate([
            {'$match': {'entNameMain': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entNameMain", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$entNameMain'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_duiwaitouzi:
            L.extend(i['hehe'])
        # print(L)
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_duiwaitouzi = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entNameMain',
                    '投资企业名称': '$entName',
                    '公司类型': '$entType',
                    '经营范围': '$business_scope',
                    '投资占比': '$funDedRatio',
                    '企业状态': '$entStatus',
                    '法人': '$name',
                    '投资金额': '$subConAm',
                    '行业': '$category',
                    '注册资金': '$regCap',
                    '省份简称': '$base',
                    '统一社会信用代码': '$creditCode',
                    '简称': '$alias',
                    '被投资企业开业日期': '$esDate'
                }}
            ])
            for data in data_duiwaitouzi:
                try:
                    writer_duiwaitouzi.writerow(data)
                    # print(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('对外投资 Over!'+ str(time.time()-start))
        f_duiwaitouzi.close()

    # 变更
    def getBiangeng(self):

        db = client['download']['ys_alter']

        f_biangeng = open(os.path.join(oriDataBusiness, '变更.csv'), 'w', encoding='utf-8', newline='')
        headers_biangeng = ['企业名称', '创建时间', '变更事项', '变更前', '变更后', '变更时间']
        writer_biangeng = csv.DictWriter(f_biangeng, headers_biangeng)
        writer_biangeng.writeheader()

        start = time.time()
        IDs_biangeng = db.aggregate([
            {'$match': {'COMPANYNAME': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$COMPANYNAME", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$COMPANYNAME'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_biangeng:
            L.extend(i['hehe'])

        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_duiwaitouzi = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$COMPANYNAME',
                    '创建时间': '$createTime',
                    '变更事项': '$altItem',
                    '变更前': '$altBe',
                    '变更后': '$altAf',
                    '变更时间': '$altDate'
                }}
            ])
            for data in data_duiwaitouzi:
                try:
                    writer_biangeng.writerow(data)
                    # print(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('变更 Over!' + str(time.time() - start))
        f_biangeng.close()

    # 分支机构
    def getFenzhijigou(self):

        db = client['download']['ys_filiation']

        f_fenzhijigou = open(os.path.join(oriDataBusiness, '分支机构.csv'), 'w', encoding='utf-8', newline='')
        headers_fenzhijigou = ['企业名称', '分支机构名称', '开业时间', '企业状态', '省份简称', '简称', '法人', '注册资金']
        writer_fenzhijigou = csv.DictWriter(f_fenzhijigou, headers_fenzhijigou)
        writer_fenzhijigou.writeheader()

        start = time.time()
        IDs_fenzhijigou = db.aggregate([
            {'$match': {'COMPANYNAME': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$COMPANYNAME", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$COMPANYNAME'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_fenzhijigou:
            L.extend(i['hehe'])

        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_fenzhijigou = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$COMPANYNAME',
                    '分支机构名称': '$brName',
                    '开业时间': '$estiblishTime',
                    '企业状态': '$regStatus',
                    '省份简称': '$base',
                    '简称': '$alias',
                    '法人': '$legalPersonName',
                    '注册资金': '$regCapital',
                }}
            ])
            for data in data_fenzhijigou:
                try:
                    if data['开业时间']:
                        if str(data['开业时间']).startswith('-'):
                            data['开业时间'] = None
                        else:
                            data['开业时间'] = datetime.datetime.fromtimestamp(data['开业时间']/1000)\
                                                .strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['开业时间'] = None
                writer_fenzhijigou.writerow(data)

            current += iter_size

        print('分支机构 Over!' + str(time.time() - start))
        f_fenzhijigou.close()

    # 经营异常
    def getJingyingyichang(self):

        db = client['download']['ys_abnormal']

        f_jingyingyichang = open(os.path.join(oriDataBusiness, '经营异常.csv'), 'w', encoding='utf-8', newline='')
        headers_jingyingyichang = ['企业名称', '列入日期', '创建时间', '列入异常名录原因', '移除异常名录原因', '作出决定机关']
        writer_jingyingyichang = csv.DictWriter(f_jingyingyichang, headers_jingyingyichang)
        writer_jingyingyichang.writeheader()

        start = time.time()
        IDs_jingyingyichang = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = [id for i in IDs_jingyingyichang for id in i['hehe']]
        # for i in IDs_jingyingyichang:
        #     L.extend(i['hehe'])
        # print(L)
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_jingyingyichang = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '列入日期': '$dateIn',  # 13位 或 10位
                    '创建时间': '$createTime',  # 13位 或 10位
                    '列入异常名录原因': '$resultIn',
                    '移除异常名录原因': '$resultOut',
                    '作出决定机关': '$orgNameOut',
                }}
            ])
            for data in data_jingyingyichang:
                try:
                    if '列入日期' in data:
                        if len(str(data['列入日期'])) > 10:
                            # data['列入日期'] = datetime.datetime.fromtimestamp(int(str(data['列入日期'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                            data['列入日期'] = datetime.datetime.fromtimestamp(data['列入日期']/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['列入日期'] = None
                try:
                    if '创建时间' in data:
                        if len(str(data['创建时间'])) > 10:
                            # data['创建时间'] = datetime.datetime.fromtimestamp(int(str(data['创建时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                            data['创建时间'] = datetime.datetime.fromtimestamp(data['创建时间']/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['创建时间'] = None
                try:
                    writer_jingyingyichang.writerow(data)
                except Exception as e:
                    print('插入失败:%s' % str(e))

            current += iter_size

        print('经营异常 Over!' + str(time.time() - start))
        f_jingyingyichang.close()

    # 行政处罚
    def getXingzhengchufa(self):

        db = client['download']['ys_caseinfo']

        f_xingzhengchufa = open(os.path.join(oriDataBusiness, '行政处罚.csv'), 'w', encoding='utf-8', newline='')
        headers_xingzhengchufa = ['企业名单', '行政处罚决定书文号', '作出行政处罚决定机关名称', '省份简称', '作出行政处罚决定日期',
                                  '行政处罚内容', '法定代表人', '注册号', '违法行为类型']
        writer_xingzhengchufa = csv.DictWriter(f_xingzhengchufa, headers_xingzhengchufa)
        writer_xingzhengchufa.writeheader()

        start = time.time()
        IDs_xingzhengchufa = db.aggregate([
            {'$match': {'COMPANYNAME': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "COMPANYNAME", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名单': {'$first': '$COMPANYNAME'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名单",
                        'storage': {'$first': '$storage'},
                        '企业名单': {'$first': '$企业名单'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_xingzhengchufa:
            L.extend(i['hehe'])

        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_xingzhengchufa = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名单': '$COMPANYNAME',
                    '行政处罚决定书文号': '$punishNumber',
                    '作出行政处罚决定机关名称': '$penAuth',
                    '省份简称': '$base',
                    '作出行政处罚决定日期': '$pendecissDate',
                    '行政处罚内容': '$penResult',
                    '法定代表人': '$legalPersonName',
                    '注册号': '$regNum',
                    '违法行为类型': '$type',
                }}
            ])
            for data in data_xingzhengchufa:
                try:
                    writer_xingzhengchufa.writerow(data)
                    # print(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('行政处罚 Over!' + str(time.time() - start))
        f_xingzhengchufa.close()

    # 董监高
    def getDongjiangao(self):

        db = client['download']['corp_lkl_person']
        f_dongjiangao = open(os.path.join(oriDataBusiness, '董监高.csv'), 'w', encoding='utf-8', newline='')
        headers_dongjiangao = ['企业名称', '姓名', '职位']
        writer_dongjiangao = csv.DictWriter(f_dongjiangao, headers_dongjiangao)
        writer_dongjiangao.writeheader()

        start = time.time()
        IDs_dongjiangao = db.aggregate([
            {'$match': {'entNameMain': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entNameMain", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entNameMain'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_dongjiangao:
            L.extend(i['hehe'])

        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_dongjiangao = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entNameMain',
                    '姓名': '$perName',
                    '职位': '$position',
                }}
            ])
            for data in data_dongjiangao:
                try:
                    writer_dongjiangao.writerow(data)
                    # print(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('董监高 Over!' + str(time.time() - start))
        f_dongjiangao.close()

    # 股东
    def getGudong(self):

        db = client['download']['corp_lkl_shholder']

        f_gudong = open(os.path.join(oriDataBusiness, '股东.csv'), 'w', encoding='utf-8', newline='')
        headers_gudong = ['企业名称', '股东名', '拥有公司个数', '认缴出资额', '调取时间']
        writer_gudong = csv.DictWriter(f_gudong, headers_gudong)
        writer_gudong.writeheader()

        regex = re.compile(r'\s+')
        start = time.time()
        IDs_gudong = db.aggregate([
            {'$match': {'entNameMain': {'$in': self.companys},
                        'storage_time': {'$not': regex}}},
            {'$group': {'_id': {'cc': "$entNameMain", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entNameMain'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = []
        for i in IDs_gudong:
            L.extend(i['hehe'])

        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_gudong = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entNameMain',
                    '股东名': '$shaName',
                    '拥有公司个数': '$toco',
                    '认缴出资额': '$subConAm',
                    '调取时间': '$storage_time',
                }}
            ])
            for data in data_gudong:
                try:
                    if data['调取时间']:
                        # data['调取时间'] = datetime.datetime.fromtimestamp(int(str(data['调取时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                        data['调取时间'] = datetime.datetime.fromtimestamp(data['调取时间']/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['调取时间'] = None
                writer_gudong.writerow(data)

            current += iter_size

        print('股东 Over!' + str(time.time() - start))
        f_gudong.close()

    # 严重违法
    def getYanzhongweifa(self):

        db = client['download']['tyc_illegalinfo']

        f_yanzhongweifa = open(os.path.join(oriDataBusiness, '严重违法.csv'), 'w', encoding='utf-8', newline='')
        headers_yanzhongweifa = ['企业名称', '列入日期', '列入原因', '决定列入部门', '移除原因', '决定移除部门']
        writer_yanzhongweifa = csv.DictWriter(f_yanzhongweifa, headers_yanzhongweifa)
        writer_yanzhongweifa.writeheader()

        start = time.time()
        IDs_yanzhongweifa = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = [id for i in IDs_yanzhongweifa for id in i['hehe']]
        # print(L)
        # for i in IDs_yanzhongweifa:
        #     L.extend(i['hehe'])
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_gudong = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '列入日期': '$putDate',
                    '列入原因': '$putReason',
                    '决定列入部门': '$putDepartment',
                    '移除原因': '$removeReason',
                    '决定移除部门': '$removeDepartment',
                }}
            ])
            for data in data_gudong:
                try:
                    if data['列入日期']:
                        # data['列入日期'] = datetime.datetime.fromtimestamp(int(str(data['列入日期'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                        data['列入日期'] = datetime.datetime.fromtimestamp(data['列入日期']/1000).strftime(
                            "%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['列入日期'] = None
                writer_yanzhongweifa.writerow(data)

            current += iter_size

        print('严重违法 Over!' + str(time.time() - start))
        f_yanzhongweifa.close()

    def run(self):
        try:
            self.getJibenxinxi()
        except Exception as e:
            print(e)
        try:
            self.getDuiwaitouzi()
        except Exception as e:
            print(e)
        try:
            self.getBiangeng()
        except Exception as e:
            print(e)
        try:
            self.getFenzhijigou()
        except Exception as e:
            print(e)
        try:
            self.getXingzhengchufa()
        except Exception as e:
            print(e)
        try:
            self.getJingyingyichang()
        except Exception as e:
            print(e)
        try:
            self.getYanzhongweifa()
        except Exception as e:
            print(e)
        try:
            self.getDongjiangao()
        except Exception as e:
            print(e)
        try:
            self.getGudong()
        except Exception as e:
            print(e)


class GetAllCourtCSV:

    def __init__(self, file):
        self.companys = file

    # 法院公告
    def get_fayuangonggao(self):

        db = client['download']['tyc_courtAnnouncement']
        f_fayuangonggao = open(os.path.join(oriDataCourt, '法院公告.csv'), 'w', encoding='utf-8', newline='')
        headers_fayuangonggao = ['企业名称', '公告id', '公告号', '公告状态号', '公告类型', '公告类型名称', '案件号', '案件内容',
                                 '法院名', '处理等级', '处理等级名字', '法官', '法官电话', '手机号', '原告', '当事人', '省份',
                                 '刊登日期', '刊登版面', '原因']
        writer_fayuangonggao = csv.DictWriter(f_fayuangonggao, headers_fayuangonggao)
        writer_fayuangonggao.writeheader()

        start = time.time()
        IDs_fayuangonggao = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])

        L = [id for i in IDs_fayuangonggao for id in i['hehe']]
        # print(L)
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_fayuangonggao = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                # {'$lookup': {'from': 'tyc_courtAnnouncement_companyList',
                #             'localField': '_id',
                #             'foreignField': 'foreignKey',
                #             'as': 'join_id'}
                #  },
                # {'$unwind': {'path': '$join_id', 'preserveNullAndEmptyArrays': True}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '公告id': '$announce_id',
                    '公告号': '$bltnno',
                    '公告状态号': '$bltnstate',
                    '公告类型': '$bltntype',
                    '公告类型名称': '$bltntypename',
                    '案件号': '$caseno',
                    '案件内容': '$content',
                    '法院名': '$courtcode',
                    '处理等级': '$dealgrade',
                    '处理等级名字': '$dealgradename',
                    '法官': '$judge',
                    '法官电话': '$judgephone',
                    '手机号': '$mobilephone',
                    '原告': '$party1',
                    '当事人': '$party2',
                    '省份': '$province',
                    '刊登日期': '$publishdate',
                    '刊登版面': '$publishpage',
                    '原因': '$reason',
                    # 'id': '$join_id'
                }}
            ])
            for data in data_fayuangonggao:
                try:
                    # if 'id' in data:
                    #     data['id'] = data['id']['id']
                    writer_fayuangonggao.writerow(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('法院公告 Over!' + str(time.time()-start))
        f_fayuangonggao.close()

    # 失信
    def get_shixin(self):

        db = client['download']['tyc_dishonest']

        f_shixin = open(os.path.join(oriDataCourt, '失信.csv'), 'w', encoding='utf-8', newline='')
        headers_shixin = ['企业名称', '失信人名称', '法人、负责人姓名', '执行依据文号', '省份地区', '身份证号码/组织机构代码',
                          '官网是否存在', '法院', '发布时间', '失信被执行人行为具体情形', '做出执行的依据单位', '生效法律文书确定的义务',
                          '立案时间', '案号']
        # dict contains fields not in fieldnames: '立案时间', '生效法律文书确定的义务', '执行依据文号'
        writer_shixin = csv.DictWriter(f_shixin, headers_shixin)
        writer_shixin.writeheader()

        start = time.time()
        IDs_shixin = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])

        L = [id for i in IDs_shixin for id in i['hehe']]
        # print(L)
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_shixin = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '失信人名称': '$iname',
                    '法人、负责人姓名': '$businessentity',
                    '执行依据文号': '$gistid',
                    '省份地区': '$areaname',
                    '身份证号码/组织机构代码': '$cardnum',
                    '官网是否存在': '$status',
                    '法院': '$courtname',
                    '发布时间': '$publishdate',
                    '失信被执行人行为具体情形': '$disruptTypeName',
                    '做出执行的依据单位': '$gistunit',
                    '生效法律文书确定的义务': '$duty',
                    '立案时间': '$regdate',
                    '案号': '$casecode',
                }}
            ])
            for data in data_shixin:
                try:
                    if data['发布时间']:
                        # data['发布时间'] = datetime.datetime.fromtimestamp(int(str(data['发布时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                        data['发布时间'] = datetime.datetime.fromtimestamp(int(data['发布时间'])/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['发布时间'] = None
                try:
                    if data['立案时间']:
                        # data['立案时间'] = datetime.datetime.fromtimestamp(int(str(data['立案时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                        data['立案时间'] = datetime.datetime.fromtimestamp(int(data['立案时间'])/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['立案时间'] = None
                writer_shixin.writerow(data)
            current += iter_size
        print('失信 Over!' + str(time.time()-start))
        f_shixin.close()

    # 开庭公告
    def get_kaitinggonggao(self):
        start = time.time()
        IDs_kaitinggonggao = db.tyc_ktannouncement.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'id': {'$first': '$_id'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'},
                        '开庭日期': {'$first': '$startDate'},
                        '法庭': {'$first': '$courtroom'},
                        '案由': {'$first': '$caseReason'},
                        '法院': {'$first': '$court'},
                        '当事人': {'$first': '$litigant'},
                        '审判长/主审人': {'$first': '$judge'},
                        '承办部门': {'$first': '$contractors'},
                        '案号': {'$first': '$caseNo'},
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'id': {'$first': '$id'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'},
                        '开庭日期': {'$first': '$开庭日期'},
                        '法庭': {'$first': '$法庭'},
                        '案由': {'$first': '$案由'},
                        '法院': {'$first': '$法院'},
                        '当事人': {'$first': '$当事人'},
                        '审判长/主审人': {'$first': '$审判长/主审人'},
                        '承办部门': {'$first': '$承办部门'},
                        '案号': {'$first': '$案号'},
                        }},
        ])
        L = [id for i in IDs_kaitinggonggao for id in i['hehe']]
        print(len(L))
        companyName = db.tyc_ktannouncement.find({'_id': {'$in': L}})
        yuangao = db.tyc_kt_plaintiff.aggregate([
            {'$match': {'foreignKey': {'$in': L}}},
            {'$group': {'_id': '$foreignKey',
                        '原告': {'$addToSet': '$name'}}}
        ])
        beigao = db.tyc_kt_defendant.aggregate([
            {'$match': {'foreignKey': {'$in': L}}},
            {'$group': {'_id': '$foreignKey',
                        '被告': {'$addToSet': '$name'}}}
        ])
        datas_main = pd.DataFrame(list(companyName))
        datas_yuangao = pd.DataFrame(list(yuangao))
        datas_beigao = pd.DataFrame(list(beigao))

        def __change_mark(llist):
            kt = '、'.join([i if llist else '' for i in llist])
            return kt

        datas_yuangao['原告'] = datas_yuangao['原告'].apply(__change_mark)
        datas_beigao['被告'] = datas_beigao['被告'].apply(__change_mark)
        datas_mix = pd.merge(datas_main, datas_yuangao, how='left', on='_id')
        datas_mix = pd.merge(datas_mix, datas_beigao, how='left', on='_id')

        def __change_date(number):
            try:
                if number:
                    # data['开庭日期'] = datetime.datetime.fromtimestamp(int(str(data['开庭日期'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                    number = datetime.datetime.fromtimestamp(int(number) / 1000).strftime("%Y-%m-%d %H:%M:%S")
            except Exception as e:
                print(number, end=', ')
                print(e)
                number = None
            return number

        datas_mix.rename(columns={'entName': '企业名称', 'startDate': '开庭日期', 'courtroom': '法庭',
                                  'caseReason': '案由', 'court': '法院', 'litigant': '当事人', 'judge': '审判长/主审人',
                                  'contractors': '承办部门', 'caseNo': '案号'}, inplace=True)
        datas_mix['开庭日期'] = datas_mix['开庭日期'].apply(__change_date)
        datas_mix.drop(['id', '_id', 'connList', 'storage_time'], axis=1, inplace=True)
        datas_mix.to_csv(os.path.join(oriDataCourt, '开庭公告_取最新无分组.csv'), index=0, encoding='utf8')

        print('开庭公告 Over!' + str(time.time() - start))

    # 欠税公告
    def get_qianshuigonggao(self):

        db = client['download']['tyc_ownTax']
        f_qianshuigonggao = open(os.path.join(oriDataCourt, '欠税公告.csv'), 'w', encoding='utf-8', newline='')
        headers_qianshuigonggao = ['企业名称', '证件号码', '法人', '经营地点', '欠税金额', '部门', '纳税人识别号',
                                   '税务类型', '纳税人类型', '欠税税种', '当前新发生欠税余额', '欠税余额', '纳税人名称',
                                   '法人证件名称', '注册类型', '发布时间']
        writer_qianshuigonggao = csv.DictWriter(f_qianshuigonggao, headers_qianshuigonggao)
        writer_qianshuigonggao.writeheader()

        start = time.time()
        IDs_qianshuigonggao = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$max': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])

        L = [id for i in IDs_qianshuigonggao for id in i['hehe']]
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            db.update_many()
            data_qianshuigonggao = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '证件号码': '$personIdNumber',
                    '法人': '$legalpersonName',
                    '经营地点': '$location',
                    '欠税金额': '$ownTaxAmount',
                    '部门': '$department',
                    '纳税人识别号': '$taxIdNumber ',
                    '税务类型': '$type',
                    '纳税人类型': '$taxpayerType',
                    '欠税税种': '$taxCategory',
                    '当前新发生欠税余额': '$newOwnTaxBalance',
                    '欠税余额': '$ownTaxBalance',
                    '纳税人名称': '$name',
                    '法人证件名称': '$personIdName',
                    '注册类型': '$regType',
                    '发布时间': '$publishDate',
                }}
            ])
            for data in data_qianshuigonggao:
                try:
                    writer_qianshuigonggao.writerow(data)
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size

        print('欠税公告 Over!' + str(time.time() - start))
        f_qianshuigonggao.close()

    # 信用中国
    def get_xinyongzhongguo(self):

        db = client['download']['tyc_creditChina']

        f_xinyongzhongguo = open(os.path.join(oriDataCourt, '信用中国.csv'), 'w', encoding='utf-8', newline='')
        headers_xinyongzhongguo = ['entName', '地区', 'url']
        writer_xinyongzhongguo = csv.DictWriter(f_xinyongzhongguo, headers_xinyongzhongguo)
        writer_xinyongzhongguo.writeheader()

        start = time.time()
        # IDs_xinyongzhongguo = db.aggregate([
        #     {'$match': {'entName': {'$in': self.companys}}},
        #     {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
        #                 'storage': {'$max': '$storage_time'},
        #                 'entName': {'$first': '$entName'},
        #                 'hehe': {'$addToSet': '$_id'}
        #                 }},
        #     {'$sort': {'storage': -1}},
        #     {'$group': {'_id': "$entName",
        #                 'storage': {'$first': '$storage'},
        #                 'entName': {'$first': '$entName'},
        #                 'hehe': {'$first': '$hehe'}
        #                 }},
        # ])
        # L = [id for i in IDs_xinyongzhongguo for id in i['hehe']]
        # list_length = len(L)
        # print(list_length)
        iter_size = 2000
        current = 0
        count = 0
        while current < len(self.companys):
            end = current + iter_size
            datas_segment = self.companys[current:end]
            data_xinyongzhongguo = db.aggregate([
                {'$match': {'entName': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    'entName': 1,
                    '地区': '$areaName', 
                    'url': 1,
                }}
            ])
            for data in data_xinyongzhongguo:
                try:
                    writer_xinyongzhongguo.writerow(data)
                    count += 1
                except Exception as e:
                    print(e)
                    print(data)
            current += iter_size
        print(count)
        print('信用中国 Over!' + str(time.time() - start))
        f_xinyongzhongguo.close()

    # 产品信息
    def get_chanpinxinxi(self):

        db = client['download']['tyc_appbkInfo']

        f_chanpinxinxi = open(os.path.join(oriDataCourt, '产品信息.csv'), 'w', encoding='utf-8', newline='')
        headers_chanpinxinxi = ['企业名称', '产品名', '领域', '产品简称', '图标', '产品分类', '描述']
        writer_chanpinxinxi = csv.DictWriter(f_chanpinxinxi, headers_chanpinxinxi)
        writer_chanpinxinxi.writeheader()

        start = time.time()
        # IDs_chanpinxinxi = db.aggregate([
        #     {'$match': {'entName': {'$in': self.companys}}},
        #     {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
        #                 'storage': {'$first': '$storage_time'},
        #                 '企业名称': {'$first': '$entName'},
        #                 'hehe': {'$addToSet': '$_id'}
        #                 }},
        #     {'$sort': {'storage': -1}},
        #     {'$group': {'_id': "$企业名称",
        #                 'storage': {'$first': '$storage'},
        #                 '企业名称': {'$first': '$企业名称'},
        #                 'hehe': {'$first': '$hehe'}
        #                 }},
        # ])
        # L = [id for i in IDs_chanpinxinxi for id in i['hehe']]
        # list_length = len(L)
        # print(list_length)
        iter_size = 2000
        current = 0
        count = 0
        while current < len(self.companys):
            end = current + iter_size
            datas_segment = self.companys[current:end]
            data_chanpinxinxi = db.aggregate([
                {'$match': {'entName': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '产品名': '$name',
                    '领域': '$classes',
                    '产品简称': '$filterName',
                    '图标': '$icon',
                    '产品分类': '$type',
                    '描述': '$brief',
                }}
            ])
            for data in data_chanpinxinxi:
                try:
                    writer_chanpinxinxi.writerow(data)
                    count += 1
                except Exception as e:
                    print(e)
                    print(data)

            current += iter_size
        print(count)
        print('产品信息 Over!' + str(time.time() - start))
        f_chanpinxinxi.close()

    # 被执行人, 暂不使用
    def get_beizhixingren(self):

        db = client['download']['tyc_zhixinginfo']

        f_beizhixingren = open(os.path.join(oriDataCourt, '被执行人.csv'), 'w', encoding='utf-8', newline='')
        headers_beizhixingren = ['企业名称', '案号', '执行法院', '被执行人名称', '身份证号／组织机构代码', '创建时间', '执行标的']
        writer_beizhixingren = csv.DictWriter(f_beizhixingren, headers_beizhixingren)
        writer_beizhixingren.writeheader()

        start = time.time()
        IDs_beizhixingren = db.aggregate([
            {'$match': {'entName': {'$in': self.companys}}},
            {'$group': {'_id': {'cc': "$entName", 'tt': '$storage_time'},
                        'storage': {'$first': '$storage_time'},
                        '企业名称': {'$first': '$entName'},
                        'hehe': {'$addToSet': '$_id'}
                        }},
            {'$sort': {'storage': -1}},
            {'$group': {'_id': "$企业名称",
                        'storage': {'$first': '$storage'},
                        '企业名称': {'$first': '$企业名称'},
                        'hehe': {'$first': '$hehe'}
                        }},
        ])
        L = [id for i in IDs_beizhixingren for id in i['hehe']]
        list_length = len(L)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = L[current:end]
            data_beizhixingren = db.aggregate([
                {'$match': {'_id': {'$in': datas_segment}}},
                {'$project': {
                    '_id': 0,
                    '企业名称': '$entName',
                    '案号': '$caseCode',
                    '执行法院': '$execCourtName',
                    '被执行人名称': '$pname',
                    '身份证号／组织机构代码': '$partyCardNum',
                    '创建时间': '$caseCreateTime',
                    '执行标的': '$execMoney',
                }}
            ])
            for data in data_beizhixingren:
                try:
                    if data['创建时间']:
                        # data['创建时间'] = datetime.datetime.fromtimestamp(int(str(data['创建时间'])[:10])).strftime("%Y-%m-%d %H:%M:%S")
                        data['创建时间'] = datetime.datetime.fromtimestamp(data['创建时间']/1000).strftime("%Y-%m-%d %H:%M:%S")
                except Exception as e:
                    print(e)
                    print(data)
                    data['创建时间'] = None
                writer_beizhixingren.writerow(data)
            current += iter_size
        print('被执行人 Over!' + str(time.time() - start))
        f_beizhixingren.close()

    def run(self):
        try:
            self.get_fayuangonggao()
        except Exception as e:
            print(e)
        try:
            self.get_shixin()
        except Exception as e:
            print(e)
        try:
            self.get_kaitinggonggao()
        except Exception as e:
            print(e)
        try:
            self.get_qianshuigonggao()
        except Exception as e:
            print(e)
        try:
            self.get_xinyongzhongguo()
        except Exception as e:
            print(e)
        try:
            self.get_chanpinxinxi()
        except Exception as e:
            print(e)


class GetAllOperation(object):

    def __init__(self, file):
        self.companys = file

    def get_dongchandiya(self):  # 动产抵押

        f_dongchandiya = open(os.path.join(oriDataOperation, '动产抵押.csv'), 'w', encoding='utf-8', newline='')
        headers_dongchandiya = ['登记机关', '登记编号', '登记日期',	'状态',	'抵押权人名称',	'抵押权人证照/证件类型',
                                '证照/证件号码',	'被担保债权种类',	'被担保债权数额',	'债务人履行债务的期限',
                                '担保的范围',	'抵押物名称',	'抵押物所有权归属',	'抵押物详细情况',	'省份',	'公示日期',
                                '注销日期',	'注销原因',	'入库时间']
        writer_dongchandiya = csv.DictWriter(f_dongchandiya, headers_dongchandiya)
        writer_dongchandiya.writeheader()

        baseInfo = db['tyc_mortgage_base']
        pawnInfoList = db['tyc_mortgage_pawn']
        peopleInfo = db['tyc_mortgage_people']
        changeInfoList = db['tyc_mortgage_change']

        start = time.time()
        list_length = len(self.companys)
        print(list_length)
        iter_size = 2000
        current = 0
        while current < list_length:
            end = current + iter_size
            datas_segment = self.companys[current:end]
            data_baseInfo = pd.DataFrame(list(baseInfo.find({'entName': {'$in': datas_segment}})))
            # data_changeInfoList = changeInfoList.find({'entName': {'$in': datas_segment}}) todo 暂时不关联该表
            data_pawnInfoList = pd.DataFrame(list(pawnInfoList.find({'entName': {'$in': datas_segment}})))
            data_peopleInfo = pd.DataFrame(list(peopleInfo.find({'entName': {'$in': datas_segment}})))

            data_dongchandiya = pd.merge(data_baseInfo, data_peopleInfo, how='left', left_on='_id', right_on='foreign_key')
            data_dongchandiya = pd.merge(data_dongchandiya, data_pawnInfoList, how='left', on='foreign_key')
            data_dongchandiya = data_dongchandiya[['regDepartment', 'regNum',	'regDate',	'status',	'peopleName',
                                                   'liceseType',	'licenseNum',	'type',	'amount',	'term',	'scope',
                                                   'pawnName',	'ownership',	'detail',	'base',	'publishDate',
                                                   'cancelDate',	'cancelReason',	'storage_time']]
            data_dongchandiya.columns = headers_dongchandiya
            for i in range(len(data_dongchandiya)):
                writer_dongchandiya.writerow(data_dongchandiya.loc[i].to_dict())
            current += iter_size
        print('动产抵押 Over!' + str(time.time() - start))
        f_dongchandiya.close()


if __name__ == '__main__':

    tt = time.time()
    # adver_wb = load_workbook(os.path.join(oriDataPath, '需要调取的名单.xlsx'))
    adver_wb = load_workbook(os.path.join(rootPath, '动产抵押调取名单.xlsx'))
    adver_ws = adver_wb.active
    companys = [i.value for i in adver_ws['A']][1:]

    # 测试专用公司名
    # company = ['新余万初投资管理中心（有限合伙）', '新余鸿蒙文化投资管理中心（有限合伙）',
    #            '江西省博越国际汽车贸易有限公司', '北京陆机科技有限公司', '北京易尚中平商贸有限公司']

    # business = GetAllBusinessCSV(companys)
    # business.run()
    # court = GetAllCourtCSV(companys)
    # court.run()
    # GetAdData(companys)
    operation = GetAllOperation(companys)
    operation.get_dongchandiya()

    adver_wb.close()
    print(time.time()-tt)
