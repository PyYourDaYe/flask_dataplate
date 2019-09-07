from flask import Flask, url_for, redirect, render_template, request, flash, \
    send_from_directory, send_file, session, Blueprint
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook
from io import BytesIO
import io
import csv
from datetime import timedelta
from scripts.基本信息 import GetAllBusinessCSV

UPLOAD_FOLDER = './uploads'
ALLOWED_EXTENSIONS = set(['xlsx', 'xls'])


app = Flask(__name__)
# bp = Blueprint('page', __name__, url_prefix='/page')
# app.register_blueprint(bp)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = '123456'
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = timedelta(seconds=1)
app.config.companylist = ''
app.config.filename = ''
# app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 限制文件大小


@app.route('/')
def index_page():
    return '<h1>index_page</h1>'


@app.route('/<username>')
def hello_world(username):
    return 'Hello World! Waiting for the perfect.'+'\n'+'you delivered %s' % username


@app.route('/url/demo1')
def demo1():
    """
        url_for是构建url，返回合成的url；
        redirect是重定向；
    """
    return redirect(url_for('hello_world', username='hehe'), 404)


'----------------------------------------------------------------------------'


@app.route('/hehe/')
@app.route('/<name>')
def demo_page1(name=None):
    return render_template('demo_page1.html', name=name)


def allowed_file(filename):  # 检测文件名
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route('/upload_file', methods=['GET', 'POST'])  # 上传文件
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(url_for('demo_page1'))
        f = request.files['file']
        # if user does not select file,browser also submit an empty part without filename
        if f.filename == '':
            flash('No selected file')
            return redirect(url_for('demo_page1'))
        if f and allowed_file(f.filename):
            # f.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename(f.filename)))

            wb = load_workbook(filename=BytesIO(f.read()))
            ws = wb.active
            companylist = [i.value for i in ws['A']]
            # flash(str(companylist))
            # return companylist
            # app.config.companylist = companylist  # todo 修改为由session存全局变量
            # app.config.filename = f.filename
            session.clear()
            session["companylist"] = companylist
            session["filename"] = f.filename
            # f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
            return render_template('demo_page1.html', companies=companylist)
            # return redirect(url_for('demo_page1'))  # 重定向本地
            # return redirect(url_for('uploaded_file', filename=f.filename)) # 反馈文件
        else:
            flash('文件格式错误')
            return redirect(url_for('demo_page1'))

    return render_template('demo_page1.html')


@app.route('/upload_text', methods=['GET', 'POST'])  # 输入公司名
def upload_text():
    if request.method == 'POST':
        companies = request.form['body']
        flash(companies)
        # if user does not select file,browser also submit an empty part without filename
        if companies == '':
            flash('No text filled')
            return redirect(url_for('demo_page1'))
        else:
            # flash(str(companylist))
            # return companylist
            # app.config.companylist = companylist  # todo 修改为由session存全局变量
            # app.config.filename = f.filename
            session.clear()
            session["companylist"] = companies
            session["filename"] = ''
            # f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
            return render_template('demo_page1.html', companies=companies)
            # return redirect(url_for('demo_page1'))  # 重定向本地
            # return redirect(url_for('uploaded_file', filename=f.filename)) # 反馈文件

    return render_template('demo_page1.html')


@app.route('/uploads/<filename>')  # 文件上传成功后的反馈
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/download_file/')
def download_file():
    # if app.config.companylist:
    companylist = session.get('companylist')
    filename = session.get('filename')
    if companylist:
        # app.config.hehe = ''
        # return send_file(r'C:\Users\zhd\Desktop\flask_hehe\project\uploads\持牌机构广告主.xlsx')
        getBusiness = GetAllBusinessCSV(companylist)
        jibenxinxi = getBusiness.getJibenxinxi()
        bf = io.BytesIO()
        jibenxinxi.to_excel(bf, index=0)
        bf.seek(0)
        if filename:
            return send_file(bf, as_attachment=True, attachment_filename='反馈基本信息-%s' % filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            return send_file(bf, as_attachment=True, attachment_filename='反馈-基本信息',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    flash('未传入公司列表')  # todo redirect的话flash没用
    return redirect(url_for('demo_page1'))


if __name__ == '__main__':
    app.run(host="0.0.0.0", port=5000, debug=True)
