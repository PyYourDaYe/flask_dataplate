from flask import url_for, redirect, render_template, flash, \
    send_from_directory, send_file, session, Blueprint, request, current_app
from manage import app
from io import BytesIO
from scripts.funcs import GetCorrespondingFile
from openpyxl import load_workbook
import os

download = Blueprint('download', __name__, url_prefix='/download')
with app.app_context():
    root_path = current_app.config['UPLOAD_FOLDER']


def get_companylist(ws):
    headerlist = [i.value for i in ws[1]]
    print(headerlist)
    if '企业名称' in headerlist:
        indx = headerlist.index('企业名称') + 1  # 列的下标:从1开始
        rowcount = ws.max_row
        # companylist = [ws.cell(row + 2, indx).value for row in range(rowcount - 1)]
        companylist = []
        for row in range(rowcount - 1):
            name = ws.cell(row + 2, indx).value
            if name is not None:
                companylist.append(name)
    else:
        companylist = [i.value for i in ws['A'] if i.value is not None]

    return companylist


@download.route('/jibenxinxi')
def jibenxinxi():
    companylist = session.pop('companies', None)
    filename = session.pop('filename', None)
    if companylist != [''] and companylist is not None:
        get_jibenxinxi = GetCorrespondingFile.getJibenxinxi(companys=companylist)
        bf = BytesIO()
        get_jibenxinxi.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈-基本信息.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif filename != '' and filename is not None:
        file = os.path.join(root_path, filename)
        wb = load_workbook(file)
        ws = wb.active
        companylist = get_companylist(ws)

        # 读取完毕删除文件并清理session
        ls = os.listdir(root_path)
        for i in ls:
            f_path = os.path.join(root_path, i)
            os.remove(f_path)
        session.clear()

        if companylist != [''] and companylist is not None:
            get_jibenxinxi = GetCorrespondingFile.getJibenxinxi(companys=companylist)
            bf = BytesIO()
            get_jibenxinxi.to_excel(bf, index=0)
            bf.seek(0)
            return send_file(bf, as_attachment=True, attachment_filename='反馈基本信息-%s' % filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    flash('未传入公司列表')  # redirect的话flash没用
    return render_template('flask首页.html')
    # return redirect(url_for('demo_page1'))
    # return send_file(r'C:\Users\zhd\Desktop\flask_hehe\project\uploads\持牌机构广告主.xlsx')


@download.route('/smokingindex')
def smokingindex():
    companylist = session.pop('companies', None)
    pca = session.pop('pca', None)
    address = session.pop('address', None)
    filename = session.pop('filename', None)
    # companylist = request.cookies.get('companylist')  # 通过cookie保存名单
    # filename = request.cookies.get('filename')
    # companylist = current_app.config.pop('companylist', None)
    # filename = current_app.config.pop("filename", None)
    if pca and address:
        print(pca, address)
        pcadict = {pca: address}
        get_smoking = GetCorrespondingFile.getSmokingindex(pcadict)
        bf = BytesIO()
        get_smoking.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈冒烟指数-%s地区.xlsx' % address,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif companylist != [''] and companylist is not None:
        get_smoking = GetCorrespondingFile.getSmokingindex(companylist)
        bf = BytesIO()
        get_smoking.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈-冒烟指数.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif filename != '' and filename is not None:
        file = os.path.join(root_path, filename)
        wb = load_workbook(file)
        ws = wb.active
        companylist = get_companylist(ws)

        # 读取完毕删除文件并清理session
        ls = os.listdir(root_path)
        for i in ls:
            f_path = os.path.join(root_path, i)
            os.remove(f_path)
        session.clear()

        if companylist != [''] and companylist is not None:
            get_smoking = GetCorrespondingFile.getSmokingindex(companylist)
            bf = BytesIO()
            get_smoking.to_excel(bf, index=0)
            bf.seek(0)
            return send_file(bf, as_attachment=True, attachment_filename='反馈冒烟指数-%s' % filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    flash('未传入公司列表')  # redirect的话flash没用
    return render_template('flask首页.html')


@download.route('/advertisement')
def advertisement():
    companylist = session.pop('companies', None)
    pca = session.pop('pca', None)
    address = session.pop('address', None)
    filename = session.pop('filename', None)
    if pca and address:
        print(pca, address)
        pcadict = {pca: address}
        ad_datas = GetCorrespondingFile.getAdv(pcadict)
        bf = BytesIO()
        ad_datas.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈广告数据-%s地区.xlsx' % address,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif companylist != [''] and companylist is not None:
        ad_datas = GetCorrespondingFile.getAdv(companylist)
        bf = BytesIO()
        ad_datas.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈-广告数据.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif filename != '' and filename is not None:
        file = os.path.join(root_path, filename)
        wb = load_workbook(file)
        ws = wb.active
        companylist = get_companylist(ws)

        # 读取完毕删除文件并清理session
        ls = os.listdir(root_path)
        for i in ls:
            f_path = os.path.join(root_path, i)
            os.remove(f_path)
        session.clear()

        if companylist != [''] and companylist is not None:
            ad_datas = GetCorrespondingFile.getAdv(companylist)
            bf = BytesIO()
            ad_datas.to_excel(bf, index=0)
            bf.seek(0)
            return send_file(bf, as_attachment=True, attachment_filename='反馈广告数据-%s' % filename,
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    flash('未传入公司列表')  # redirect的话flash没用
    return render_template('flask首页.html')


@download.route('/fenlei')
def fenlei():
    companylist = session.pop('companies', None)
    filename = session.pop('filename', None)

    # sheet_ob = current_app.config.pop("sheet_ob", None)
    # if sheet_ob is not None:
    #     sheet = sheet_ob.values
    #     get_fenlei = GetCorrespondingFile.company_sort(sheet)
    if companylist != [''] and companylist is not None:
        get_fenlei = GetCorrespondingFile.company_sort(companylist)
        bf = BytesIO()
        get_fenlei.to_excel(bf, index=0)
        bf.seek(0)
        return send_file(bf, as_attachment=True, attachment_filename='反馈-公司分类.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    elif filename != '' and filename is not None:
        file = os.path.join(root_path, filename)
        wb = load_workbook(file)
        ws = wb.active
        get_fenlei = GetCorrespondingFile.company_sort(ws.values)

        # 读取完毕删除文件并清理session
        ls = os.listdir(root_path)
        for i in ls:
            f_path = os.path.join(root_path, i)
            os.remove(f_path)
        session.clear()

        bf = BytesIO()
        get_fenlei.to_excel(bf, index=0)
        bf.seek(0)

        return send_file(bf, as_attachment=True, attachment_filename='反馈公司分类-%s' % filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    else:
        flash('未传入公司列表')
        return render_template('flask首页.html')
