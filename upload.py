from flask import url_for, redirect, render_template, request, flash, \
    send_from_directory, send_file, session, Blueprint, current_app, make_response
from io import BytesIO
from openpyxl import load_workbook
import pandas as pd

upload = Blueprint('upload', __name__, url_prefix='/upload')
ALLOWED_EXTENSIONS = set(['xlsx', 'xls'])


def allowed_file(filename):  # 检测文件名
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# todo 对上传的名单做数据清洗
@upload.route('/upload_file', methods=['GET', 'POST'])  # 上传文件
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part')
            return redirect(url_for('dataplate_index'))
        f = request.files['file']
        # if user does not select file,browser also submit an empty part without filename
        if f.filename == '':
            flash('No selected file')
            return redirect(url_for('dataplate_index'))
        if f and allowed_file(f.filename):
            wb = load_workbook(filename=BytesIO(f.read()))
            temp = pd.DataFrame(BytesIO(f.read()))
            print(temp.head())
            ws = wb.active
            headerlist = [i.value for i in ws[1]]
            print(headerlist)
            if '企业名称' in headerlist:
                indx = headerlist.index('企业名称')+1  # 列的下标
                rowcount = ws.max_row
                companylist = [ws.cell(row+2, indx).value for row in range(rowcount-1)]
            else:
                companylist = [i.value for i in ws['A']]
            # print(companylist)

            # session.clear()  #
            # session["companylist"] = companylist
            session["filename"] = f.filename

                  # f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))  # 保存上传的文件到服务器
            current_app.config['companylist'] = companylist
            # current_app.config["filename"] = f.filename
            current_app.config['sheet_ob'] = ws
            return render_template('flask首页.html', companies=companylist)

            # resp = make_response(render_template('flask首页.html', companies=companylist))
            # resp.set_cookie('companylist', str(companylist))
            # resp.set_cookie('filename', f.filename)
            # return resp

            # return redirect(url_for('demo_page1'))  # 重定向本地
            # return redirect(url_for('uploaded_file', filename=f.filename)) # 反馈文件
        else:
            flash('文件格式错误')
            return redirect(url_for('dataplate_index'))

    return render_template('flask首页.html')


@upload.route('/upload_text', methods=['GET', 'POST'])  # 输入公司名
def upload_text():
    if request.method == 'POST':
        # temp = request.form['text']
        temp = request.form['input_text']
        companies = temp.split('\r\n')
        print(companies)
        # if user does not select file,browser also submit an empty part without filename
        if companies == ['']:
            flash('No text filled')
            return redirect(url_for('dataplate_index'))
        else:
            # flash(str(companylist))
            # return companylist
            current_app.config['companylist'] = companies
            # session.clear()
            # session["companylist"] = companies
            # session["filename"] = ''
            # f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
            return render_template('flask首页.html', companies=companies)
            # return redirect(url_for('demo_page1'))  # 重定向本地
            # return redirect(url_for('uploaded_file', filename=f.filename)) # 反馈文件

    return render_template('flask首页.html')


@upload.route('/upload_pca', methods=['GET', 'POST'])  # 输入省市区 province\city\cityarea
def upload_pca():
    if request.method == 'POST':
        pca = request.values.get('pca')
        # pca = request.args.get('pca')
        address = request.form['address_name']
        if address and pca:
            address = str(address).strip().replace('\r\n', '')
            print(pca)
            print(address)
            session.clear()
            session['pca'] = pca
            session['address'] = address
            # f.save(os.path.join(app.config['UPLOAD_FOLDER'], f.filename))
            # flash(address)  # todo 以其他形式在页面反馈
            return render_template('flask首页.html', address=address)
            # return redirect(url_for('dataplate_index'))  # 重定向本地
        else:
            flash('No text filled or selected')
            return redirect(url_for('dataplate_index'))

    return render_template('flask首页.html')


@upload.route('/uploads/<filename>')  # 文件上传成功后的反馈
def uploaded_file(filename):
    return send_from_directory(current_app.config['UPLOAD_FOLDER'], filename)
